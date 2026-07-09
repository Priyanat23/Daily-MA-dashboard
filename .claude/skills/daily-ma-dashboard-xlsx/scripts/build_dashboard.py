# -*- coding: utf-8 -*-
"""
Daily MA Dashboard -> Excel builder.

This script rebuilds the Daily MA Dashboard (normally produced as HTML by the
`daily-ma-dashboard-v2` skill) as a fully formula-driven .xlsx workbook with
3 sheets: Dashboard, DailyMA, RawData.

CURRENT FIELD SET (8 fields, do not silently drop back to 7):
jrts, ap, ar, rc, so1, so2, pr, sv  -- Shop Online is split into so1/so2 to
match the HTML dashboard exactly. All column positions (target_total,
target_pct, actual_total, actual_pct, daily_revenue, and every downstream
Dashboard/DailyMA column) are DERIVED from len(FIELDS) via the N_FIELDS /
TGT_*/ACT_*/DM_*/COMP_* variables defined right after RawData is built --
never hardcode a column letter for "the total column" etc. If FIELDS ever
changes length again, everything reflows automatically EXCEPT the two spots
noted in "Key rules" below (breakdown card grouping assumes a multiple of 4).

KPI CARDS: 5 cards across the top of Dashboard (Incentive, Gross MA รายเดือน,
Daily Gross MA วันล่าสุด, %MA/Revenue วันล่าสุด, %MA/Revenue สะสมเดือน) using
column widths A:C/D:F/G:I/J:L/M:P. Incentive is NOT read from RawData daily
rows -- it comes from separate INCENTIVE_TARGET/INCENTIVE_ACTUAL config values
(see CONFIG block) stored in RawData as their own config rows + IncentiveTarget/
IncentiveActual defined names, because the source Google Sheet keeps Incentive
in `1.Dashboard` cells C3/C4, not in the daily `2.DailyMA` rows.

CROSS-MONTH PERIODS: the reporting period does not always start on day 1 of a
calendar month (e.g. week-based periods like "29 Jun - 26 Jul" spanning two
months). Never build display text by concatenating a raw day number with a
single fixed MONTH_LABEL -- it breaks the moment the period crosses a month
boundary. Always pull the actual date string for a given dayNum from
RawData!C via INDEX/MATCH instead (see how the header/banner/comp-table labels
do this). DATE_RANGE_SHORT exists only because the narrow comp-table column B
can't fit a full "D month YYYY - D month YYYY" string; keep it short and
year-less.

HOW TO USE THIS SCRIPT
-----------------------
1. Edit ONLY the "CONFIG: EDIT HERE" block below with the current period's data
   (pulled from the same Google Sheet FILE_ID used by daily-ma-dashboard-v2).
2. Run:  python3 build_dashboard.py
3. Recalculate formulas and check for errors:
   python3 /mnt/skills/public/xlsx/scripts/recalc.py daily_ma_dashboard.xlsx 60
   -> total_errors MUST be 0. If not, fix the offending formula and rerun.
4. Copy the output file to /mnt/user-data/outputs/ and call present_files.

Do NOT change anything below the "END CONFIG" marker unless you specifically
need to change the dashboard's structure/formulas/charts (these were built to
mirror the HTML dashboard's layout, formulas, and conditional-formatting
rules 1:1; see SKILL.md "Key rules" section before touching them).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import FormulaRule

# =====================================================================
# CONFIG: EDIT HERE — update every month / whenever the source sheet changes
# =====================================================================

OUTPUT_PATH = "/home/claude/daily_ma_dashboard.xlsx"

DEPT_NAME = "PCRC"
WEEK_LABEL = "27–30"                                     # ช่วงสัปดาห์ปัจจุบัน (ตามชื่อในชีต)
DATE_RANGE_LABEL = "29 มิถุนายน – 26 กรกฎาคม 2569"        # ช่วงวันที่เต็มของรอบสัปดาห์ 27–30 (ข้ามเดือน)
MONTH_LABEL = "มิถุนายน–กรกฎาคม 2569"                     # ใช้เป็นข้อความอ้างอิงทั่วไปเท่านั้น (วันที่จริงอ่านจาก RawData!C แทน เพราะช่วงข้อมูลคาบเกี่ยว 2 เดือน)
DATE_RANGE_SHORT = "29 มิ.ย.–26 ก.ค."                      # ช่วงวันที่แบบย่อ (ไม่มีปี) สำหรับคอลัมน์แคบ เช่น comp table
COMPARE_LABEL = "สัปดาห์ 10–13 (มี.ค. 2569)"                # ป้ายชื่อสัปดาห์ที่ใช้เทียบ

TODAY_DAY = 3           # ลำดับวันที่ปัจจุบัน (1 ก.ค. 2569) ในช่วงข้อมูล 29 มิ.ย.–26 ก.ค. (29 มิ.ย.=1, 30 มิ.ย.=2, 1 ก.ค.=3)
REVENUE_MONTH = 5927657.111   # Revenue ของสายธุรกิจ (บาท)
INCENTIVE_TARGET = 26.36813187   # Incentive ประจำวัน (บาท) - Target (จากชีต 1.Dashboard เซลล์ C3)
INCENTIVE_ACTUAL = 941.23        # Incentive ประจำวัน (บาท) - Actual (จากชีต 1.Dashboard เซลล์ C4)

# รายได้ของวันที่ทราบ revenue เจาะจง (ปกติคือวันล่าสุด/เมื่อวาน) -> {dayNum: revenue}
DAILY_REVENUE = {2: 211702.0397}

# สัปดาห์ -> ช่วงวันที่แบบสั้น (ใช้ในตารางสรุปรายสัปดาห์ของ Dashboard)
WEEKS = [(27, "29 มิ.ย.–5 ก.ค."), (28, "6–12 ก.ค."), (29, "13–19 ก.ค."), (30, "20–26 ก.ค.")]

# ข้อมูลรายวันทั้งหมด (1 entry ต่อวัน, ครบทุกวันของช่วงสัปดาห์ 27–30: 29 มิ.ย.–26 ก.ค. 2569)
# รูปแบบ: (dayNum, week, 'D mmm. YYYY', holiday_bool, target_dict, actual_dict)  -- 'date' รวมปี พ.ศ. เพราะช่วงข้อมูลคาบเกี่ยว 2 เดือน
# target/actual dict keys: jrts, ap, ar, rc, so1, so2, pr, sv, pct  (so1/so2 = Shop Online 1/2 แยกกัน, total = SUM ของ 8 ฟิลด์แรก, คำนวณเป็นสูตรใน Excel)
DAYS = [
 (1,27,'29 มิ.ย. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=-20851.79,ap=0,ar=0,rc=0,so1=-114.59,so2=371.36,pr=0,sv=0,pct=-0.097283)),
 (2,27,'30 มิ.ย. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=-24862.39,ap=0,ar=0,rc=0,so1=-70.13,so2=846.75,pr=0,sv=0,pct=-0.113772)),
 (3,27,'1 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (4,27,'2 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (5,27,'3 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (6,27,'4 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (7,27,'5 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (8,28,'6 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (9,28,'7 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (10,28,'8 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (11,28,'9 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (12,28,'10 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (13,28,'11 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (14,28,'12 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (15,29,'13 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (16,29,'14 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (17,29,'15 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (18,29,'16 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (19,29,'17 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (20,29,'18 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (21,29,'19 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000296),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (22,30,'20 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (23,30,'21 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (24,30,'22 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (25,30,'23 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (26,30,'24 ก.ค. 2569',False, dict(jrts=-24273.8649,ap=76.79918,ar=3.846154,rc=6.704117,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=-0.114196),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (27,30,'25 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
 (28,30,'26 ก.ค. 2569',True, dict(jrts=0,ap=76.79918,ar=0,rc=0,so1=4.748597,so2=6.198584,pr=0,sv=0,pct=0.000414),
                         dict(jrts=0,ap=0,ar=0,rc=0,so1=0,so2=0,pr=0,sv=0,pct=0)),
]

# ข้อมูลสัปดาห์อ้างอิง (ช่วงเทียบ เช่นเดือนก่อนหรือปีก่อน) — Actual ของช่วงนั้น (hardcode เพราะเป็นชุดข้อมูลคนละช่วงเวลา)
REF_COMPARISON = dict(
    label="Actual สัปดาห์ 10–13 (มี.ค. 2569)",
    period="2–29 มี.ค. 2569",
    values=[-535443.1143, 3231.941368, 0, 0, 0, 0, 0, 0],   # jrts, ap, ar, rc, so1, so2, pr, sv
)

# ข้อความวิเคราะห์ผล 3 คอลัมน์ — แต่ละคอลัมน์เป็น list ของ bullet (string)
ANALYSIS_TEXT = dict(
    cause=[
        "JRTS สะสม 2 วันแรก (29–30 มิ.ย.) -45,714.18 บาท เป็นต้นทุนแรงงานหลักที่ดึง MA ลง คิดเป็น 102.3% ของ MA ติดลบทั้งหมด",
        "Association Product, Refer, Reciprocation ยังไม่มี Actual เลยตลอด 2 วันแรกของสัปดาห์ 27 (0 บาท จาก Target รวม 174.70 บาท)",
        "Shop Online (1+2) สะสม 1,033.39 บาท ดีกว่า Target 21.89 บาท อย่างมีนัยสำคัญ ช่วยพยุง MA รวมไว้",
    ],
    source=[
        "แผนก PCRC มีต้นทุนแรงงาน JRTS สูงตามโครงสร้างพนักงาน ทำให้ Gross MA ติดลบต่อเนื่องทุกวันเป็นปกติ",
        "ข้อมูล Association ยังไม่ถูกบันทึกลงระบบ หรือยังไม่มีการโทรติดตามลูกค้าในช่วง 2 วันแรกของสัปดาห์ 27",
        "%MA/Revenue สะสม -0.75% ดีกว่า Target -8.15% มาก เนื่องจาก JRTS Actual ต่ำกว่า Target และ Shop Online ทำได้เกิน Target",
    ],
    solution=[
        "เร่งบันทึก Association Product/Refer/Reciprocation ที่ตกค้าง และเริ่มโทรติดตามลูกค้าให้ครบ Target รวม ~87.35 บาท/วัน",
        "รักษาระดับ Shop Online ต่อเนื่อง ปัจจุบันทำได้เกิน Target กว่า 46 เท่า ให้คงอัตรานี้ตลอดสัปดาห์",
        "เฝ้าติดตาม %MA/Revenue วันถัดไปให้ยังคงดีกว่า Target ต่อเนื่อง เนื่องจากยังเหลืออีก 26 วันในรอบสัปดาห์ 27–30",
    ],
)

# =====================================================================
# END CONFIG — ไม่ต้องแก้โค้ดด้านล่างนี้ตามปกติ
# =====================================================================

BLUE_DARK, BLUE_TXT, BLUE_FILL = "1E40AF", "1E3A8A", "DBEAFE"
HEADER_FILL_1, HEADER_FILL_2 = "DBEAFE", "EDE9FE"
GREEN_TXT, RED_TXT, GRAY_TXT, DARK_TXT = "16A34A", "DC2626", "6B7280", "1F2937"
GREEN_FILL, RED_FILL, GRAY_FILL, HOLIDAY_FILL = "DCFCE7", "FEE2E2", "F1F5F9", "FEF9EC"
WHITE, FONT_NAME = "FFFFFF", "Arial"

thin = Side(style="thin", color="DBEAFE")
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(size=10, bold=False, color=DARK_TXT, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FIELDS = ["jrts", "ap", "ar", "rc", "so1", "so2", "pr", "sv"]
FIELD_LABELS = ["JRTS", "Association Product", "Association Refer", "Reciprocation",
                "Shop Online 1", "Shop Online 2", "Project", "Service"]

def merge_set(ws, rng, value, font=None, fill_color=None, align=CENTER, number_format=None, border=False):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    cell = ws[first]
    cell.value = value
    cell.font = font if font else F()
    if fill_color:
        for row in ws[rng]:
            for c in row:
                c.fill = fill(fill_color)
    cell.alignment = align
    if number_format:
        cell.number_format = number_format
    if border:
        for row in ws[rng]:
            for c in row:
                c.border = BORDER_ALL
    return cell

def color_code(ws, cell_addr, formula_gt, formula_lt):
    ws.conditional_formatting.add(cell_addr, FormulaRule(formula=[formula_gt], font=Font(name=FONT_NAME, color=GREEN_TXT, bold=True)))
    ws.conditional_formatting.add(cell_addr, FormulaRule(formula=[formula_lt], font=Font(name=FONT_NAME, color=RED_TXT, bold=True)))

wb = Workbook()

# ====================================================================
# SHEET: RawData
# ====================================================================
rd = wb.active
rd.title = "RawData"
rd.sheet_view.showGridLines = False

headers = ["dayNum", "week", "date", "holiday",
           "Target JRTS", "Target Asso.Product", "Target Asso.Refer", "Target Reciprocation",
           "Target Shop Online 1", "Target Shop Online 2", "Target Project", "Target Service",
           "target_total", "Target %MA",
           "Actual JRTS", "Actual Asso.Product", "Actual Asso.Refer", "Actual Reciprocation",
           "Actual Shop Online 1", "Actual Shop Online 2", "Actual Project", "Actual Service",
           "actual_total", "Actual %MA",
           "daily_revenue"]
N_FIELDS = len(FIELDS)                                  # 8
TGT_START, TGT_END = 5, 4 + N_FIELDS                     # E..L
TGT_TOTAL_COL, TGT_PCT_COL = TGT_END + 1, TGT_END + 2     # M, N
ACT_START, ACT_END = TGT_PCT_COL + 1, TGT_PCT_COL + N_FIELDS  # O..V
ACT_TOTAL_COL, ACT_PCT_COL = ACT_END + 1, ACT_END + 2     # W, X
REV_COL = ACT_PCT_COL + 1                                 # Y
N_COLS = REV_COL

for ci, h in enumerate(headers, start=1):
    c = rd.cell(row=1, column=ci, value=h)
    c.font = F(9, True, WHITE); c.fill = fill(BLUE_DARK); c.alignment = CENTER

TGT_TOTAL_LTR, TGT_PCT_LTR = get_column_letter(TGT_TOTAL_COL), get_column_letter(TGT_PCT_COL)
ACT_TOTAL_LTR, ACT_PCT_LTR = get_column_letter(ACT_TOTAL_COL), get_column_letter(ACT_PCT_COL)
TGT_START_LTR, TGT_END_LTR = get_column_letter(TGT_START), get_column_letter(TGT_END)
ACT_START_LTR, ACT_END_LTR = get_column_letter(ACT_START), get_column_letter(ACT_END)
REV_LTR = get_column_letter(REV_COL)

for ri, (dnum, week, date, holiday, tgt, act) in enumerate(DAYS, start=2):
    rd.cell(row=ri, column=1, value=dnum)
    rd.cell(row=ri, column=2, value=week)
    rd.cell(row=ri, column=3, value=date)
    rd.cell(row=ri, column=4, value=holiday)
    for fi, f in enumerate(FIELDS):
        rd.cell(row=ri, column=TGT_START+fi, value=tgt[f]).font = F(9, color="0000FF")
    rd.cell(row=ri, column=TGT_TOTAL_COL, value=f"=SUM({TGT_START_LTR}{ri}:{TGT_END_LTR}{ri})").font = F(9, bold=True)
    rd.cell(row=ri, column=TGT_PCT_COL, value=tgt["pct"]).number_format = "0.00%"
    rd.cell(row=ri, column=TGT_PCT_COL).font = F(9, color="0000FF")
    for fi, f in enumerate(FIELDS):
        rd.cell(row=ri, column=ACT_START+fi, value=act[f]).font = F(9, color="0000FF")
    rd.cell(row=ri, column=ACT_TOTAL_COL, value=f"=SUM({ACT_START_LTR}{ri}:{ACT_END_LTR}{ri})").font = F(9, bold=True)
    rd.cell(row=ri, column=ACT_PCT_COL, value=act["pct"]).number_format = "0.00%"
    rd.cell(row=ri, column=ACT_PCT_COL).font = F(9, color="0000FF")
    if dnum in DAILY_REVENUE:
        rd.cell(row=ri, column=REV_COL, value=DAILY_REVENUE[dnum]).font = F(9, color="0000FF")
    if holiday:
        for col in range(1, N_COLS+1):
            rd.cell(row=ri, column=col).fill = fill(HOLIDAY_FILL)

for col, w in zip(range(1, N_COLS+1), [7, 6, 14, 7] + [10]*(N_COLS-4)):
    rd.column_dimensions[get_column_letter(col)].width = w

n_days = len(DAYS)
last_row = 1 + n_days
rd.cell(row=last_row+2, column=2, value="Config").font = F(11, True, BLUE_DARK)
rd.cell(row=last_row+3, column=2, value="TodayDay (วันที่ปัจจุบัน 1-28)").font = F(10, True)
rd.cell(row=last_row+3, column=3, value=TODAY_DAY).font = F(10, True, color="0000FF")
rd.cell(row=last_row+4, column=2, value="RevenueMonth (บาท)").font = F(10, True)
rd.cell(row=last_row+4, column=3, value=REVENUE_MONTH).font = F(10, True, color="0000FF")
rd.cell(row=last_row+4, column=3).number_format = "#,##0.00"
rd.cell(row=last_row+5, column=2, value="Incentive Target ประจำวัน (บาท)").font = F(10, True)
rd.cell(row=last_row+5, column=3, value=INCENTIVE_TARGET).font = F(10, True, color="0000FF")
rd.cell(row=last_row+5, column=3).number_format = "#,##0.00"
rd.cell(row=last_row+6, column=2, value="Incentive Actual ประจำวัน (บาท)").font = F(10, True)
rd.cell(row=last_row+6, column=3, value=INCENTIVE_ACTUAL).font = F(10, True, color="0000FF")
rd.cell(row=last_row+6, column=3).number_format = "#,##0.00"
rd.cell(row=last_row+7, column=2,
        value="หมายเหตุ: เปลี่ยน TodayDay เพื่อเลื่อนวันที่ใช้คำนวณ Breakdown / KPI วันล่าสุด").font = F(8, False, GRAY_TXT, italic=True)

TODAY_CELL = f"$C${last_row+3}"
REV_CELL = f"$C${last_row+4}"
INC_TGT_CELL = f"$C${last_row+5}"
INC_ACT_CELL = f"$C${last_row+6}"
wb.defined_names["TodayDay"] = DefinedName("TodayDay", attr_text=f"RawData!{TODAY_CELL}")
wb.defined_names["RevenueMonth"] = DefinedName("RevenueMonth", attr_text=f"RawData!{REV_CELL}")
wb.defined_names["IncentiveTarget"] = DefinedName("IncentiveTarget", attr_text=f"RawData!{INC_TGT_CELL}")
wb.defined_names["IncentiveActual"] = DefinedName("IncentiveActual", attr_text=f"RawData!{INC_ACT_CELL}")

RD_LAST = last_row  # last data row in RawData (row 1 + n_days)
target_cols = [get_column_letter(TGT_START+i) for i in range(N_FIELDS)]
actual_cols = [get_column_letter(ACT_START+i) for i in range(N_FIELDS)]

print(f"RawData built ({n_days} days, rows 2-{RD_LAST})")

# ====================================================================
# SHEET: Dashboard
# ====================================================================
ds = wb.create_sheet("Dashboard")
ds.sheet_view.showGridLines = False
dash_widths = {"A":32,"B":18,"C":12,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":10,"K":10,"L":13,"M":11,"N":11,"O":11,"P":11}
for col, w in dash_widths.items():
    ds.column_dimensions[col].width = w
ds.page_setup.orientation = "landscape"
ds.page_setup.fitToWidth = 1
ds.page_setup.fitToHeight = 0
ds.sheet_properties.pageSetUpPr.fitToPage = True

merge_set(ds, "A1:P1", f"📊 Daily MA Dashboard — {DEPT_NAME}", F(16, True, BLUE_DARK), HEADER_FILL_1, CENTER)
ds.row_dimensions[1].height = 28
merge_set(ds, "A2:P2",
    f'="หน่วยงาน: {DEPT_NAME}    |    สัปดาห์ที่: {WEEK_LABEL}    |    ช่วงเวลา: {DATE_RANGE_LABEL}"',
    F(10, False, BLUE_TXT), BLUE_FILL, CENTER)
merge_set(ds, "A3:P3",
    f'="ข้อมูล ณ วันที่: "&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&"    |    Revenue สายธุรกิจ: ฿"&TEXT(RevenueMonth,"#,##0.00")&'
    f'"    |    เปรียบเทียบกับ: {COMPARE_LABEL}"',
    F(10, False, BLUE_TXT), BLUE_FILL, CENTER)
ds.row_dimensions[2].height = 18
ds.row_dimensions[3].height = 18

merge_set(ds, "A5:H5", f"📆 สะสม ({DATE_RANGE_LABEL}) สัปดาห์ {WEEK_LABEL}", F(12, True, BLUE_DARK), WHITE, CENTER, border=True)
merge_set(ds, "I5:P5", f'="📅 ล่าสุด วันที่ "&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&" (สัปดาห์ "&INDEX(RawData!B2:B{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&")"',
          F(12, True, BLUE_DARK), WHITE, CENTER, border=True)

period_rows = [
    ("Revenue ของสายธุรกิจ", "REVENUE"),
    ("Gross MA Actual", "ACT_TOTAL"),
    ("Gross MA Target", "TGT_TOTAL"),
    ("Gap (Actual − Target)", "GAP"),
    ("%MA/Revenue Actual", "PCT_ACT"),
    ("%MA/Revenue Target", "PCT_TGT"),
]
for i, (label, kind) in enumerate(period_rows):
    r = 6 + i
    numfmt = "#,##0.00" if kind == "REVENUE" else ("0.00%" if kind in ("PCT_ACT", "PCT_TGT") else "#,##0.00;(#,##0.00)")
    if kind == "GAP":
        numfmt = '"▲ +"#,##0.00;"▼ -"#,##0.00'
    merge_set(ds, f"A{r}:E{r}", label, F(10, False, GRAY_TXT), WHITE, LEFT, border=True)
    merge_set(ds, f"I{r}:M{r}", label, F(10, False, GRAY_TXT), WHITE, LEFT, border=True)
    if kind == "REVENUE":
        f1, f2 = "=RevenueMonth", f"=INDEX(RawData!{REV_LTR}2:{REV_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))"
    elif kind == "ACT_TOTAL":
        f1, f2 = f"=SUM(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST})", f"=INDEX(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))"
    elif kind == "TGT_TOTAL":
        f1, f2 = f"=SUM(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST})", f"=INDEX(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))"
    elif kind == "GAP":
        f1, f2 = "=F7-F8", "=N7-N8"
    elif kind == "PCT_ACT":
        f1, f2 = "=F7/F6", "=N7/N6"
    else:
        f1, f2 = "=F8/F6", "=N8/N6"
    merge_set(ds, f"F{r}:H{r}", f1, F(10, True, DARK_TXT), WHITE, RIGHT, number_format=numfmt, border=True)
    merge_set(ds, f"N{r}:P{r}", f2, F(10, True, DARK_TXT), WHITE, RIGHT, number_format=numfmt, border=True)

color_code(ds, "F7", "F7>F8", "F7<F8"); color_code(ds, "F10", "F10>F11", "F10<F11")
color_code(ds, "N7", "N7>N8", "N7<N8"); color_code(ds, "N10", "N10>N11", "N10<N11")
for r in range(6, 12):
    ds.row_dimensions[r].height = 16

kpi_defs = [
    ("Incentive ประจำวัน (บาท)", "=IncentiveActual", "=IncentiveTarget", "#,##0"),
    ("Gross MA รายเดือน (บาท)", f"=SUM(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST})", f"=SUM(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST})", "#,##0"),
    ("Daily Gross MA วันล่าสุด (บาท)",
     f"=INDEX(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))",
     f"=INDEX(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))", "#,##0"),
    ("%MA/Revenue วันล่าสุด",
     f"=INDEX(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))/INDEX(RawData!{REV_LTR}2:{REV_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))",
     f"=INDEX(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))/INDEX(RawData!{REV_LTR}2:{REV_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))", "0.00%"),
    ("%MA/Revenue สะสมเดือน", f"=SUM(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST})/RevenueMonth", f"=SUM(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST})/RevenueMonth", "0.00%"),
]
kpi_cols = ["A:C", "D:F", "G:I", "J:L", "M:P"]
for i, (label, fact, ftgt, numfmt) in enumerate(kpi_defs):
    colrng = kpi_cols[i]; c0 = colrng[0]
    merge_set(ds, f"{c0}13:{colrng[-1]}13", label, F(9, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
    merge_set(ds, f"{c0}14:{colrng[-1]}14", fact, F(18, True, DARK_TXT), WHITE, CENTER, number_format=numfmt, border=True)
    merge_set(ds, f"{c0}15:{colrng[-1]}15", ftgt, F(9, False, GRAY_TXT), WHITE, CENTER, number_format='"Target: "'+numfmt, border=True)
    gapf = f'=IF(({c0}14-{c0}15)>=0,"▲ + ดีกว่า Target "&TEXT(ABS({c0}14-{c0}15),"{numfmt}"),"▼ - แย่กว่า Target "&TEXT(ABS({c0}14-{c0}15),"{numfmt}"))'
    merge_set(ds, f"{c0}16:{colrng[-1]}16", gapf, F(9, True, GREEN_TXT), GREEN_FILL, CENTER, border=True)
    ds.conditional_formatting.add(f"{c0}16:{colrng[-1]}16", FormulaRule(
        formula=[f"({c0}14-{c0}15)<0"], fill=fill(RED_FILL), font=Font(name=FONT_NAME, bold=True, color=RED_TXT)))
    color_code(ds, f"{c0}14", f"{c0}14>{c0}15", f"{c0}14<{c0}15")
ds.row_dimensions[13].height = 16; ds.row_dimensions[14].height = 26
ds.row_dimensions[15].height = 16; ds.row_dimensions[16].height = 16

print("Dashboard: header/period/kpi built")

merge_set(ds, "A18:H18", "📈 Target vs Actual %MA รายวัน", F(11, True, "374151"), WHITE, CENTER)
merge_set(ds, "I18:P18", "📊 Components สะสม (Target vs Actual)", F(11, True, "374151"), WHITE, CENTER)

line1 = LineChart()
line1.title = "Target vs Actual %MA รายวัน"
line1.height, line1.width = 8, 15.5
cats = Reference(rd, min_col=3, min_row=2, max_row=RD_LAST)
line1.add_data(Reference(rd, min_col=TGT_PCT_COL, min_row=1, max_row=RD_LAST), titles_from_data=True)
line1.add_data(Reference(rd, min_col=ACT_PCT_COL, min_row=1, max_row=RD_LAST), titles_from_data=True)
line1.set_categories(cats)
line1.y_axis.numFmt = "0.00%"; line1.y_axis.title = "%MA/Revenue"
ds.add_chart(line1, "A19")

# (Bar chart + breakdown/weekly/comp table row numbers are computed below once we
#  know how many rows the breakdown table needs; for 7 fields they are fixed.)
BREAKDOWN_HDR_ROW = 39
COMP_HDR_ROW = 66       # fixed: matches the static rows used by bar chart references below
COMP_REF_ROW, COMP_TGT_ROW, COMP_ACT_ROW = 67, 68, 69

bar1 = BarChart()
bar1.type = "col"; bar1.title = "Components: Target vs Actual"
bar1.height, bar1.width = 8, 15.5
COMP_FIELD_MAXCOL = 2 + N_FIELDS  # C..J for 8 fields
cats2 = Reference(ds, min_col=3, max_col=COMP_FIELD_MAXCOL, min_row=COMP_HDR_ROW, max_row=COMP_HDR_ROW)
ser_t = Series(Reference(ds, min_col=3, max_col=COMP_FIELD_MAXCOL, min_row=COMP_TGT_ROW, max_row=COMP_TGT_ROW), title="Target (เต็มเดือน)")
ser_a = Series(Reference(ds, min_col=3, max_col=COMP_FIELD_MAXCOL, min_row=COMP_ACT_ROW, max_row=COMP_ACT_ROW), title="Actual (สะสมถึงปัจจุบัน)")
bar1.series.append(ser_t); bar1.series.append(ser_a)
bar1.set_categories(cats2)
ds.add_chart(bar1, "I19")
for r in range(19, 36):
    ds.row_dimensions[r].height = 15
print("Dashboard: charts placed")

merge_set(ds, "A38:P38", f'="Breakdown รายสายธุรกิจ (สะสม "&RawData!C2&"–"&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&")"', F(12, True, BLUE_DARK), WHITE, LEFT)

bd_groups = [FIELD_LABELS[0:4], FIELD_LABELS[4:8]]
bd_cols_groups = [list(zip(target_cols[0:4], actual_cols[0:4])), list(zip(target_cols[4:8], actual_cols[4:8]))]
bd_col_ranges = ["A:D", "E:H", "I:L", "M:P"]
for gi, (labels, colpairs) in enumerate(zip(bd_groups, bd_cols_groups)):
    r0 = BREAKDOWN_HDR_ROW + gi * 4
    rName, rAct, rTgt, rGap = r0, r0 + 1, r0 + 2, r0 + 3
    for i, (label, (tcol, acol)) in enumerate(zip(labels, colpairs)):
        colrng = bd_col_ranges[i]; c0 = colrng[0]
        merge_set(ds, f"{c0}{rName}:{colrng[-1]}{rName}", label, F(9, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
        f_act = f'=SUMIFS(RawData!{acol}2:{acol}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)'
        f_tgt = f'=SUMIFS(RawData!{tcol}2:{tcol}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)'
        merge_set(ds, f"{c0}{rAct}:{colrng[-1]}{rAct}", f_act, F(13, True, DARK_TXT), WHITE, CENTER, number_format="#,##0.00;(#,##0.00)", border=True)
        merge_set(ds, f"{c0}{rTgt}:{colrng[-1]}{rTgt}", f_tgt, F(8, False, GRAY_TXT), WHITE, CENTER, number_format='"Target: "#,##0.00;"Target: ("#,##0.00")"', border=True)
        gapf = f'=IF(({c0}{rAct}-{c0}{rTgt})>=0,"▲ +"&TEXT(ABS({c0}{rAct}-{c0}{rTgt}),"#,##0.00"),"▼ "&TEXT(ABS({c0}{rAct}-{c0}{rTgt}),"#,##0.00"))'
        merge_set(ds, f"{c0}{rGap}:{colrng[-1]}{rGap}", gapf, F(9, True, GREEN_TXT), GREEN_FILL, CENTER, border=True)
        ds.conditional_formatting.add(f"{c0}{rGap}:{colrng[-1]}{rGap}", FormulaRule(
            formula=[f"({c0}{rAct}-{c0}{rTgt})<0"], fill=fill(RED_FILL), font=Font(name=FONT_NAME, bold=True, color=RED_TXT)))
        color_code(ds, f"{c0}{rAct}", f"{c0}{rAct}>{c0}{rTgt}", f"{c0}{rAct}<{c0}{rTgt}")
print("Dashboard: breakdown built")
for gi in range(2):
    r0 = BREAKDOWN_HDR_ROW + gi * 4
    ds.row_dimensions[r0].height = 16
    ds.row_dimensions[r0+1].height = 22
    ds.row_dimensions[r0+2].height = 14
    ds.row_dimensions[r0+3].height = 16

WEEK_TBL_HDR_ROW = 50
merge_set(ds, f"A49:L49", "ตารางสรุปรายสัปดาห์", F(12, True, BLUE_DARK), WHITE, LEFT)
for ci, h in enumerate(["สัปดาห์", "ช่วงวันที่", "ประเภท"] + FIELD_LABELS + ["Total", "%MA/Rev"], start=1):
    cell = ds.cell(row=WEEK_TBL_HDR_ROW, column=ci, value=h)
    cell.font = F(8, True, WHITE); cell.fill = fill(BLUE_DARK); cell.alignment = CENTER; cell.border = BORDER_ALL

r = WEEK_TBL_HDR_ROW + 1
for wnum, wlabel in WEEKS:
    rT, rA, rG = r, r+1, r+2
    ds.cell(row=rT, column=1, value=wnum).font = F(9, True)
    ds.cell(row=rT, column=2, value=wlabel).font = F(9)
    ds.cell(row=rT, column=3, value="Target").font = F(9, False, GRAY_TXT)
    ds.cell(row=rA, column=3, value="Actual").font = F(9, False, DARK_TXT)
    ds.cell(row=rG, column=3, value="Gap").font = F(9, True)
    for fi, (tcol, acol) in enumerate(zip(target_cols + [TGT_TOTAL_LTR], actual_cols + [ACT_TOTAL_LTR])):
        col = 4 + fi
        cT = ds.cell(row=rT, column=col, value=f'=SUMIFS(RawData!{tcol}2:{tcol}{RD_LAST},RawData!B2:B{RD_LAST},{wnum})')
        cA = ds.cell(row=rA, column=col, value=f'=SUMIFS(RawData!{acol}2:{acol}{RD_LAST},RawData!B2:B{RD_LAST},{wnum})')
        cG = ds.cell(row=rG, column=col, value=f"={get_column_letter(col)}{rA}-{get_column_letter(col)}{rT}")
        for c in (cT, cA, cG):
            c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT; c.border = BORDER_ALL
        cT.font = F(9, False, GRAY_TXT); cA.font = F(9, True, DARK_TXT); cG.font = F(9, True)
    pct_col_idx = 4 + len(target_cols) + 1  # column after Total
    pct_col_ltr = get_column_letter(pct_col_idx)
    pT = ds.cell(row=rT, column=pct_col_idx, value=f'=AVERAGEIFS(RawData!{TGT_PCT_LTR}2:{TGT_PCT_LTR}{RD_LAST},RawData!B2:B{RD_LAST},{wnum})')
    pA = ds.cell(row=rA, column=pct_col_idx, value=f'=AVERAGEIFS(RawData!{ACT_PCT_LTR}2:{ACT_PCT_LTR}{RD_LAST},RawData!B2:B{RD_LAST},{wnum})')
    pG = ds.cell(row=rG, column=pct_col_idx, value=f"={pct_col_ltr}{rA}-{pct_col_ltr}{rT}")
    for c in (pT, pA, pG):
        c.number_format = "0.00%"; c.alignment = RIGHT; c.border = BORDER_ALL
    pT.font = F(9, False, GRAY_TXT); pA.font = F(9, True, DARK_TXT); pG.font = F(9, True)
    for cc in range(1, 4):
        for rr2 in (rT, rA, rG):
            ds.cell(row=rr2, column=cc).border = BORDER_ALL
        ds.cell(row=rT, column=cc).fill = fill(GRAY_FILL)
    color_code(ds, f"D{rA}:L{rA}", f"D{rA}>D{rT}", f"D{rA}<D{rT}")
    color_code(ds, f"{pct_col_ltr}{rA}", f"{pct_col_ltr}{rA}>{pct_col_ltr}{rT}", f"{pct_col_ltr}{rA}<{pct_col_ltr}{rT}")
    r += 3
print("Dashboard: weekly table built")

merge_set(ds, "A65:P65", f"เปรียบเทียบกับสัปดาห์อ้างอิง ({COMPARE_LABEL})", F(12, True, BLUE_DARK), WHITE, LEFT)
for ci, h in enumerate(["รายการ", "ช่วงเวลา"] + FIELD_LABELS + ["Total", "%MA/Rev"], start=1):
    cell = ds.cell(row=COMP_HDR_ROW, column=ci, value=h)
    cell.font = F(8, True, WHITE); cell.fill = fill(BLUE_DARK); cell.alignment = CENTER; cell.border = BORDER_ALL

COMP_FIELD_END_LTR = get_column_letter(2 + N_FIELDS)   # J for 8 fields
COMP_TOTAL_COL = 3 + N_FIELDS                          # K
COMP_PCT_COL = COMP_TOTAL_COL + 1                      # L
COMP_TOTAL_LTR = get_column_letter(COMP_TOTAL_COL)
COMP_PCT_LTR = get_column_letter(COMP_PCT_COL)

ds.cell(row=COMP_REF_ROW, column=1, value=REF_COMPARISON["label"]).font = F(9)
ds.cell(row=COMP_REF_ROW, column=2, value=REF_COMPARISON["period"]).font = F(9)
for fi, v in enumerate(REF_COMPARISON["values"]):
    c = ds.cell(row=COMP_REF_ROW, column=3+fi, value=v); c.font = F(9, color="0000FF")
    c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT
ds.cell(row=COMP_REF_ROW, column=COMP_TOTAL_COL, value=f"=SUM(C{COMP_REF_ROW}:{COMP_FIELD_END_LTR}{COMP_REF_ROW})").number_format = "#,##0.00;(#,##0.00)"
ds.cell(row=COMP_REF_ROW, column=COMP_PCT_COL, value=f"={COMP_TOTAL_LTR}{COMP_REF_ROW}/RevenueMonth").number_format = "0.00%"

ds.cell(row=COMP_TGT_ROW, column=1, value=f"Target สัปดาห์ {WEEK_LABEL} ({DATE_RANGE_LABEL})").font = F(9)
ds.cell(row=COMP_TGT_ROW, column=2, value=DATE_RANGE_SHORT).font = F(9)
for fi, tcol in enumerate(target_cols):
    c = ds.cell(row=COMP_TGT_ROW, column=3+fi, value=f"=SUM(RawData!{tcol}2:{tcol}{RD_LAST})")
    c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT; c.font = F(9, True)
ds.cell(row=COMP_TGT_ROW, column=COMP_TOTAL_COL, value=f"=SUM(C{COMP_TGT_ROW}:{COMP_FIELD_END_LTR}{COMP_TGT_ROW})").number_format = "#,##0.00;(#,##0.00)"
ds.cell(row=COMP_TGT_ROW, column=COMP_PCT_COL, value=f"={COMP_TOTAL_LTR}{COMP_TGT_ROW}/RevenueMonth").number_format = "0.00%"

ds.cell(row=COMP_ACT_ROW, column=1, value=f'="Actual สะสม (ถึง "&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&")"').font = F(9, True)
ds.cell(row=COMP_ACT_ROW, column=2, value=f'=LEFT(RawData!C2,LEN(RawData!C2)-5)&"–"&LEFT(INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),LEN(INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)))-5)').font = F(9)
for fi, acol in enumerate(actual_cols):
    c = ds.cell(row=COMP_ACT_ROW, column=3+fi, value=f'=SUMIFS(RawData!{acol}2:{acol}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)')
    c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT; c.font = F(9, True)
ds.cell(row=COMP_ACT_ROW, column=COMP_TOTAL_COL, value=f"=SUM(C{COMP_ACT_ROW}:{COMP_FIELD_END_LTR}{COMP_ACT_ROW})").number_format = "#,##0.00;(#,##0.00)"
ds.cell(row=COMP_ACT_ROW, column=COMP_PCT_COL, value=f"={COMP_TOTAL_LTR}{COMP_ACT_ROW}/RevenueMonth").number_format = "0.00%"

for rr in (COMP_REF_ROW, COMP_TGT_ROW, COMP_ACT_ROW):
    for cc in range(1, COMP_PCT_COL+1):
        ds.cell(row=rr, column=cc).border = BORDER_ALL
    ds.cell(row=rr, column=COMP_TOTAL_COL).font = F(9, True); ds.cell(row=rr, column=COMP_PCT_COL).font = F(9, True)
color_code(ds, f"C{COMP_ACT_ROW}:{COMP_TOTAL_LTR}{COMP_ACT_ROW}", f"C{COMP_ACT_ROW}>C{COMP_TGT_ROW}", f"C{COMP_ACT_ROW}<C{COMP_TGT_ROW}")
print("Dashboard: comp table built")

merge_set(ds, "A72:P72", "การวิเคราะห์ผล", F(12, True, BLUE_DARK), WHITE, LEFT)
merge_set(ds, "A73:E73", "🔍 สาเหตุ", F(10, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
merge_set(ds, "F73:K73", "📌 ที่มาของสาเหตุ", F(10, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
merge_set(ds, "L73:P73", "✅ แนวทางแก้ปัญหา", F(10, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
def bullets(lst):
    return "\n\n".join(f"{i+1}. {t}" for i, t in enumerate(lst))
merge_set(ds, "A74:E76", bullets(ANALYSIS_TEXT["cause"]), F(9, False, "374151"), WHITE, LEFT, border=True)
merge_set(ds, "F74:K76", bullets(ANALYSIS_TEXT["source"]), F(9, False, "374151"), WHITE, LEFT, border=True)
merge_set(ds, "L74:P76", bullets(ANALYSIS_TEXT["solution"]), F(9, False, "374151"), WHITE, LEFT, border=True)
for rr in range(74, 77):
    ds.row_dimensions[rr].height = 30
print("Dashboard: analysis built")

banner_fmla = (
    f'=IF(INDEX(RawData!{ACT_PCT_LTR}2:{ACT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))>=INDEX(RawData!{TGT_PCT_LTR}2:{TGT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),'
    f'"✅ %MA/Revenue เมื่อวาน ("&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&") = "&TEXT(INDEX(RawData!{ACT_PCT_LTR}2:{ACT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),"0.00%")&'
    f'" ดีกว่า Target ("&TEXT(INDEX(RawData!{TGT_PCT_LTR}2:{TGT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),"0.00%")&") — ไม่มีการมอบหมายงานเพิ่มเติม",'
    f'"⚠️ %MA/Revenue เมื่อวาน ("&INDEX(RawData!C2:C{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0))&") = "&TEXT(INDEX(RawData!{ACT_PCT_LTR}2:{ACT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),"0.00%")&'
    f'" ต่ำกว่า Target ("&TEXT(INDEX(RawData!{TGT_PCT_LTR}2:{TGT_PCT_LTR}{RD_LAST},MATCH(TodayDay-1,RawData!A2:A{RD_LAST},0)),"0.00%")&") — ต้องมอบหมายงานเพิ่มเติม ตรวจสอบทีม ap/ar/rc/so1/so2")'
)
merge_set(ds, "A79:P79", banner_fmla, F(11, True, "166534"), GREEN_FILL, CENTER, border=True)
ds.row_dimensions[79].height = 24
ds.conditional_formatting.add("A79:P79", FormulaRule(
    formula=['LEFT($A$79,1)="⚠"'], fill=fill(RED_FILL), font=Font(name=FONT_NAME, bold=True, color="991B1B")))
print("Dashboard: assignment banner built")

# ====================================================================
# SHEET: DailyMA
# ====================================================================
dm = wb.create_sheet("DailyMA")
dm.sheet_view.showGridLines = False
DM_FIELD_START = 4
DM_TOTAL_COL = DM_FIELD_START + N_FIELDS      # L (12) for 8 fields
DM_PCT_COL = DM_TOTAL_COL + 1                 # M (13)
DM_TOTAL_LTR, DM_PCT_LTR = get_column_letter(DM_TOTAL_COL), get_column_letter(DM_PCT_COL)
dm_widths = [7, 16, 9] + [10]*N_FIELDS + [11, 9]
for i, w in enumerate(dm_widths, start=1):
    dm.column_dimensions[get_column_letter(i)].width = w
dm.page_setup.orientation = "landscape"
dm.page_setup.fitToWidth = 1
dm.page_setup.fitToHeight = 0
dm.sheet_properties.pageSetUpPr.fitToPage = True

merge_set(dm, f"A1:{DM_PCT_LTR}1", f"📅 DailyMA — รายวัน สัปดาห์ {WEEK_LABEL} ({DATE_RANGE_LABEL})", F(16, True, BLUE_DARK), HEADER_FILL_2, CENTER)
dm.row_dimensions[1].height = 26
merge_set(dm, f"A2:{DM_PCT_LTR}2", f'="ช่วง: {DATE_RANGE_LABEL}    |    สัปดาห์: {WEEK_LABEL}    |    Revenue: ฿"&TEXT(RevenueMonth,"#,##0.00")',
          F(10, False, BLUE_TXT), BLUE_FILL, CENTER)

dm_kpi = [
    ("JRTS สะสม (Actual)", f'=SUMIFS(RawData!{ACT_START_LTR}2:{ACT_START_LTR}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)', f"=SUM(RawData!{TGT_START_LTR}2:{TGT_START_LTR}{RD_LAST})"),
    ("Shop Online สะสม (Actual)",
     f'=SUMIFS(RawData!{get_column_letter(ACT_START+4)}2:{get_column_letter(ACT_START+4)}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)+SUMIFS(RawData!{get_column_letter(ACT_START+5)}2:{get_column_letter(ACT_START+5)}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)',
     f"=SUM(RawData!{get_column_letter(TGT_START+4)}2:{get_column_letter(TGT_START+4)}{RD_LAST})+SUM(RawData!{get_column_letter(TGT_START+5)}2:{get_column_letter(TGT_START+5)}{RD_LAST})"),
    ("Total MA สะสม (Actual)", f'=SUMIFS(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)', f"=SUM(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST})"),
    ("%MA/Rev สะสม", f'=SUMIFS(RawData!{ACT_TOTAL_LTR}2:{ACT_TOTAL_LTR}{RD_LAST},RawData!A2:A{RD_LAST},"<"&TodayDay)/RevenueMonth', f"=SUM(RawData!{TGT_TOTAL_LTR}2:{TGT_TOTAL_LTR}{RD_LAST})/RevenueMonth"),
]
for i, (label, fact, ftgt) in enumerate(dm_kpi):
    colrng = ["A:C", "D:F", "G:I", "J:L"][i]; nf = "0.00%" if i == 3 else "#,##0"
    merge_set(dm, f"{colrng[0]}4:{colrng[-1]}4", label, F(9, True, BLUE_DARK), HEADER_FILL_1, CENTER, border=True)
    merge_set(dm, f"{colrng[0]}5:{colrng[-1]}5", fact, F(15, True, DARK_TXT), WHITE, CENTER, number_format=nf, border=True)
    merge_set(dm, f"{colrng[0]}6:{colrng[-1]}6", ftgt, F(9, False, GRAY_TXT), WHITE, CENTER, number_format='"Target: "'+nf, border=True)
    color_code(dm, f"{colrng[0]}5", f"{colrng[0]}5>{colrng[0]}6", f"{colrng[0]}5<{colrng[0]}6")

merge_set(dm, f"A8:{DM_PCT_LTR}8", "ตารางรายวัน Target / Actual / Gap", F(12, True, BLUE_DARK), WHITE, LEFT)
for ci, h in enumerate(["สัปดาห์", "วันที่", "ประเภท"] + FIELD_LABELS + ["Total", "%MA/Rev"], start=1):
    cell = dm.cell(row=9, column=ci, value=h)
    cell.font = F(8, True, WHITE); cell.fill = fill(BLUE_DARK); cell.alignment = CENTER; cell.border = BORDER_ALL

r = 10
week_start_row = {}
for dnum, week, date, holiday, tgt, act in DAYS:
    rd_row = dnum + 1
    rT, rA, rG = r, r+1, r+2
    week_start_row.setdefault(week, rT)
    dm.cell(row=rT, column=2, value=date + (" (วันหยุด)" if holiday else "")).font = F(8)
    dm.merge_cells(f"B{rT}:B{rG}")
    dm.cell(row=rT, column=3, value="Target").font = F(8, False, GRAY_TXT)
    dm.cell(row=rA, column=3, value="Actual").font = F(8, False, DARK_TXT)
    dm.cell(row=rG, column=3, value="Gap").font = F(8, True)
    for fi in range(N_FIELDS):
        col = DM_FIELD_START + fi
        rdcol_t, rdcol_a = get_column_letter(TGT_START+fi), get_column_letter(ACT_START+fi)
        cT = dm.cell(row=rT, column=col, value=f"=RawData!{rdcol_t}{rd_row}")
        cA = dm.cell(row=rA, column=col, value=f'=IF(RawData!A{rd_row}<TodayDay,RawData!{rdcol_a}{rd_row},"")')
        cG = dm.cell(row=rG, column=col, value=f'=IF(RawData!A{rd_row}<TodayDay,{get_column_letter(col)}{rA}-{get_column_letter(col)}{rT},"")')
        for c in (cT, cA, cG):
            c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT; c.border = BORDER_ALL
        cT.font = F(8, False, GRAY_TXT); cA.font = F(8, True, DARK_TXT); cG.font = F(8, True)
    cTt = dm.cell(row=rT, column=DM_TOTAL_COL, value=f"=RawData!{TGT_TOTAL_LTR}{rd_row}")
    cAt = dm.cell(row=rA, column=DM_TOTAL_COL, value=f'=IF(RawData!A{rd_row}<TodayDay,RawData!{ACT_TOTAL_LTR}{rd_row},"")')
    cGt = dm.cell(row=rG, column=DM_TOTAL_COL, value=f'=IF(RawData!A{rd_row}<TodayDay,{DM_TOTAL_LTR}{rA}-{DM_TOTAL_LTR}{rT},"")')
    cTp = dm.cell(row=rT, column=DM_PCT_COL, value=f"=RawData!{TGT_PCT_LTR}{rd_row}")
    cAp = dm.cell(row=rA, column=DM_PCT_COL, value=f'=IF(RawData!A{rd_row}<TodayDay,RawData!{ACT_PCT_LTR}{rd_row},"")')
    cGp = dm.cell(row=rG, column=DM_PCT_COL, value=f'=IF(RawData!A{rd_row}<TodayDay,{DM_PCT_LTR}{rA}-{DM_PCT_LTR}{rT},"")')
    for c in (cTt, cAt, cGt):
        c.number_format = "#,##0.00;(#,##0.00)"; c.alignment = RIGHT; c.border = BORDER_ALL
    for c in (cTp, cAp, cGp):
        c.number_format = "0.00%"; c.alignment = RIGHT; c.border = BORDER_ALL
    for c in (cTt, cAt, cGt, cTp, cAp, cGp):
        c.font = F(8, True)
    for cc in (1, 2, 3):
        for rr2 in (rT, rA, rG):
            dm.cell(row=rr2, column=cc).border = BORDER_ALL
    if holiday:
        for cc in range(1, DM_PCT_COL+1):
            for rr2 in (rT, rA, rG):
                dm.cell(row=rr2, column=cc).fill = fill(HOLIDAY_FILL)
    r += 3

for week, srow in week_start_row.items():
    erow = srow + 20
    dm.merge_cells(f"A{srow}:A{erow}")
    dm.cell(row=srow, column=1, value=week).font = F(9, True)
    dm.cell(row=srow, column=1).alignment = CENTER
    dm.cell(row=srow, column=1).border = BORDER_ALL

last_table_row = r - 1
print(f"DailyMA: table built, last row {last_table_row}")

chart_title_row1 = last_table_row + 2
chart_anchor1 = chart_title_row1 + 1
merge_set(dm, f"A{chart_title_row1}:{DM_PCT_LTR}{chart_title_row1}", "กราฟ Target vs Actual % รายวัน", F(11, True, "374151"), WHITE, LEFT)
line2 = LineChart()
line2.title = "Target vs Actual %MA รายวัน"
line2.height, line2.width = 9, 24
line2.add_data(Reference(rd, min_col=TGT_PCT_COL, min_row=1, max_row=RD_LAST), titles_from_data=True)
line2.add_data(Reference(rd, min_col=ACT_PCT_COL, min_row=1, max_row=RD_LAST), titles_from_data=True)
line2.set_categories(Reference(rd, min_col=3, min_row=2, max_row=RD_LAST))
line2.y_axis.numFmt = "0.00%"
dm.add_chart(line2, f"A{chart_anchor1}")

chart_title_row2 = chart_anchor1 + 19
chart_anchor2 = chart_title_row2 + 1
merge_set(dm, f"A{chart_title_row2}:{DM_PCT_LTR}{chart_title_row2}", "Components รายวัน Actual (บาท)", F(11, True, "374151"), WHITE, LEFT)
bar2 = BarChart()
bar2.type = "col"; bar2.grouping = "stacked"; bar2.overlap = 100
bar2.title = "Components รายวัน Actual"
bar2.height, bar2.width = 9, 24
bar2.add_data(Reference(rd, min_col=ACT_START, max_col=ACT_END, min_row=1, max_row=RD_LAST), titles_from_data=True)
bar2.set_categories(Reference(rd, min_col=3, min_row=2, max_row=RD_LAST))
dm.add_chart(bar2, f"A{chart_anchor2}")
print("DailyMA: charts placed")

ds.freeze_panes = "A4"
dm.freeze_panes = "A10"
wb._sheets = [wb["Dashboard"], wb["DailyMA"], wb["RawData"]]
wb.active = 0

wb.save(OUTPUT_PATH)
print("SAVED:", OUTPUT_PATH)
